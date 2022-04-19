from openpyxl import load_workbook

wb = load_workbook("Project_HomeWork\Excel\North.xlsx")

ws = wb['North']

colum_a = ws['A']
colum_b = ws['B']

row = ws[1]

# print(wb,ws)
# print(colum_a)
# print(colum_a[1].value)

for i in row : #or colum
    print(i.value)

# for i in colum_b :
#     i.value = i.value * 10
#     print(i.value)

# wb.save('new_northx10.xlsx') #save new file


