import tkinter as tk
import openpyxl
from openpyxl import load_workbook
from tkinter import *


root = tk.Tk()
wb = load_workbook('Excel1/North.xlsx')
ws = wb["North"]

colum_a = ws['A']
colum_b = ws['B']

row = ws[1]

for i in range(1,len(colum_a)+1) :
    for j in range(len(row)) :
        Label(root,text=row[j].value).grid(column=j,row=i)
    row = ws[1+i]

root.mainloop()